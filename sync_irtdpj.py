"""
sync_irtdpj.py
Varre todos os associados do IRTDPJ Brasil e exporta:
  - associados.xlsx           (cadastro)
  - detalhe_remessas.xlsx     (uma linha por mensalidade)
  - extrato_por_cartorio.xlsx (uma linha por cartório, colunas = meses)

Requer:
  pip install requests python-dateutil openpyxl

Como rodar:
  Linux/Mac:  export IRTDPJ_TOKEN="eyJhbGc..."
              python sync_irtdpj.py
  Windows:    $env:IRTDPJ_TOKEN="eyJhbGc..."
              python sync_irtdpj.py
"""

import os
import re
import sys
import time
import json
import logging
import requests
from collections import defaultdict
from pathlib import Path
from dateutil import parser as dateparser
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ===================== CONFIGURAÇÃO =====================
API_BASE = "https://gestaoapi.irtdpjbrasil.org.br"
TOKEN = os.environ.get("IRTDPJ_TOKEN")
PAUSA_ENTRE_REQS = 0.1
TIMEOUT = 30
MAX_RETRIES = 3
FRONTEND_DATA_PATH = Path(__file__).resolve().parent / "gestaocontratov1" / "public" / "data" / "irtdpj.json"

# ===================== LOGGING =====================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("irtdpj")


# ===================== HTTP =====================
def montar_sessao(token):
    if not token:
        log.error("Variável de ambiente IRTDPJ_TOKEN não definida.")
        sys.exit(1)
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    })
    return s


def request_com_retry(session, method, url, **kwargs):
    for tentativa in range(1, MAX_RETRIES + 1):
        try:
            r = session.request(method, url, timeout=TIMEOUT, **kwargs)
            if r.status_code == 401:
                log.error("Token expirado/invalido (401). Renove e rode de novo.")
                sys.exit(2)
            if r.status_code >= 500:
                raise requests.HTTPError(f"{r.status_code} server error")
            r.raise_for_status()
            return r
        except (requests.ConnectionError, requests.Timeout, requests.HTTPError) as e:
            espera = 2 ** tentativa
            log.warning(f"Falha {tentativa}/{MAX_RETRIES} ({e}). Aguardando {espera}s...")
            time.sleep(espera)
    raise RuntimeError(f"Esgotadas tentativas: {method} {url}")


def listar_associados(session):
    log.info("Buscando lista de associados...")
    r = request_com_retry(
        session, "POST", f"{API_BASE}/api/associados/busca",
        json={"descricao": "", "filtros": [], "numPage": 1,
              "numRows": 10000, "tipoAssociadoId": None},
    )
    data = r.json()
    associados = data.get("data", [])
    log.info(f"Encontrados {len(associados)} associados.")
    return associados


def buscar_detalhe(session, associado_id):
    r = request_com_retry(session, "GET", f"{API_BASE}/api/associados/{associado_id}")
    return r.json()


# ===================== UTIL =====================
def parse_data(s):
    if not s:
        return None
    try:
        return dateparser.isoparse(s)
    except (ValueError, TypeError):
        return None


def fmt_data(s):
    d = parse_data(s)
    return d.strftime("%d/%m/%Y") if d else ""


def extrair_mes_ano(descricao, data_vencimento):
    """
    Tenta extrair MM/YYYY da descricao (ex: "MENSALIDADE 03/2026").
    Se nao conseguir, usa a data de vencimento.
    Retorna ((ano, mes), "MM/YYYY") ou (None, None).
    """
    if descricao:
        m = re.search(r"(\d{1,2})/(\d{4})", descricao)
        if m:
            mes, ano = int(m.group(1)), int(m.group(2))
            if 1 <= mes <= 12:
                return (ano, mes), f"{mes:02d}/{ano}"
    d = parse_data(data_vencimento)
    if d:
        return (d.year, d.month), f"{d.month:02d}/{d.year}"
    return None, None


def extrair_associado(detalhe):
    assoc = detalhe.get("associacao") or {}
    contato = detalhe.get("contato") or {}
    return {
        "associadoId":      detalhe.get("associadoId"),
        "nome":             detalhe.get("nome"),
        "cpfCnpjPessoa":    detalhe.get("cpfCnpj"),
        "nomeCartorio":     assoc.get("nomeCartorio"),
        "cnpjCartorio":     assoc.get("cnpjCartorio"),
        "cns":              assoc.get("cns"),
        "dataAssociacao":   fmt_data(assoc.get("dataAssociacao")),
        "valorMensalidade": assoc.get("mensalidade"),
        "email":            contato.get("email"),
        "telefone":         contato.get("telefone"),
        "ativo":            detalhe.get("ativo"),
    }


# ===================== EXPORTACAO =====================
FILL_PAGO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PEND = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEAD = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEAD = Font(bold=True, color="FFFFFF")


def ajustar_largura(ws, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:300]:
            v = str(cell.value) if cell.value is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def estilizar_header(ws, num_cols):
    for i in range(1, num_cols + 1):
        c = ws.cell(row=1, column=i)
        c.fill = FILL_HEAD
        c.font = FONT_HEAD
        c.alignment = Alignment(horizontal="center", vertical="center")


def salvar_cadastro(caminho, cadastros):
    wb = Workbook()
    ws = wb.active
    ws.title = "Associados"
    headers = ["ID", "Nome", "CPF/CNPJ Pessoa", "Cartório", "CNPJ Cartório",
               "CNS", "Data Associação", "Mensalidade (R$)", "Email", "Telefone", "Ativo"]
    ws.append(headers)
    estilizar_header(ws, len(headers))
    for c in cadastros:
        ws.append([c["associadoId"], c["nome"], c["cpfCnpjPessoa"],
                   c["nomeCartorio"], c["cnpjCartorio"], c["cns"],
                   c["dataAssociacao"], c["valorMensalidade"],
                   c["email"], c["telefone"], c["ativo"]])
    ws.freeze_panes = "A2"
    ajustar_largura(ws)
    wb.save(caminho)
    log.info(f"Salvo: {caminho} ({len(cadastros)} linhas)")


def salvar_detalhe(caminho, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Remessas"
    headers = ["Cartório", "CNPJ", "CNS", "Descrição", "Tipo",
               "Valor (R$)", "Vencimento", "Data Pagamento", "Status"]
    ws.append(headers)
    estilizar_header(ws, len(headers))
    for linha in linhas:
        ws.append(linha)
        status = linha[-1]
        fill = FILL_PAGO if status == "PAGO" else FILL_PEND
        ws.cell(row=ws.max_row, column=len(headers)).fill = fill
    ws.freeze_panes = "A2"
    ajustar_largura(ws)
    wb.save(caminho)
    log.info(f"Salvo: {caminho} ({len(linhas)} linhas)")


def salvar_extrato_pivotado(caminho, cadastros, remessas_por_cartorio, meses_ordenados):
    """
    Uma linha por cartório, uma coluna por mês.
    Célula: 'PAGO 09/03' (verde) ou 'PENDENTE' (vermelho) ou vazio.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato por Cartório"

    headers = ["Cartório", "CNPJ", "CNS", "Mensalidade (R$)",
               "Total Pago (R$)", "Total Pendente (R$)", "Qtd Paga", "Qtd Pendente"]
    headers += [m_label for (_, m_label) in meses_ordenados]
    ws.append(headers)
    estilizar_header(ws, len(headers))

    cadastros_por_id = {c["associadoId"]: c for c in cadastros}
    ids_ordenados = sorted(
        cadastros_por_id.keys(),
        key=lambda i: (cadastros_por_id[i].get("nomeCartorio") or
                       cadastros_por_id[i].get("nome") or "").strip().upper()
    )

    for aid in ids_ordenados:
        c = cadastros_por_id[aid]
        remessas_dict = remessas_por_cartorio.get(aid, {})

        total_pago = 0.0
        total_pend = 0.0
        qtd_pago = 0
        qtd_pend = 0
        celulas_meses = []

        for (chave_mes, _) in meses_ordenados:
            r = remessas_dict.get(chave_mes)
            if r is None:
                celulas_meses.append(("", None))
                continue
            valor = r.get("valorRemessa") or 0
            if r.get("dataPago"):
                d = parse_data(r.get("dataPago"))
                texto = f"PAGO {d.strftime('%d/%m/%y')}" if d else "PAGO"
                celulas_meses.append((texto, "pago"))
                total_pago += valor
                qtd_pago += 1
            else:
                celulas_meses.append(("PENDENTE", "pend"))
                total_pend += valor
                qtd_pend += 1

        nome_exib = (c.get("nomeCartorio") or c.get("nome") or "").strip()
        linha_base = [
            nome_exib,
            c.get("cnpjCartorio"),
            c.get("cns"),
            c.get("valorMensalidade"),
            round(total_pago, 2),
            round(total_pend, 2),
            qtd_pago,
            qtd_pend,
        ]
        ws.append(linha_base + [t for (t, _) in celulas_meses])

        row_idx = ws.max_row
        for offset, (_, status) in enumerate(celulas_meses, start=len(linha_base) + 1):
            if status == "pago":
                ws.cell(row=row_idx, column=offset).fill = FILL_PAGO
            elif status == "pend":
                ws.cell(row=row_idx, column=offset).fill = FILL_PEND

    ws.freeze_panes = "B2"
    ajustar_largura(ws, max_width=30)
    wb.save(caminho)
    log.info(f"Salvo: {caminho}")


def salvar_frontend_json(caminho, cadastros, detalhe_linhas, remessas_por_cartorio, meses_ordenados):
    meses = [m_label for (_, m_label) in meses_ordenados]
    cadastros_por_id = {c["associadoId"]: c for c in cadastros}
    ids_ordenados = sorted(
        cadastros_por_id.keys(),
        key=lambda i: (cadastros_por_id[i].get("nomeCartorio") or
                       cadastros_por_id[i].get("nome") or "").strip().upper()
    )

    extrato = []
    for aid in ids_ordenados:
        c = cadastros_por_id[aid]
        remessas_dict = remessas_por_cartorio.get(aid, {})
        meses_status = {}
        total_pago = 0.0
        total_pendente = 0.0
        qtd_paga = 0
        qtd_pendente = 0

        for (chave_mes, label_mes) in meses_ordenados:
            r = remessas_dict.get(chave_mes)
            if not r:
                meses_status[label_mes] = ""
                continue

            valor = r.get("valorRemessa") or 0
            if r.get("dataPago"):
                d = parse_data(r.get("dataPago"))
                meses_status[label_mes] = f"PAGO {d.strftime('%d/%m/%y')}" if d else "PAGO"
                total_pago += valor
                qtd_paga += 1
            else:
                meses_status[label_mes] = "PENDENTE"
                total_pendente += valor
                qtd_pendente += 1

        extrato.append({
            "cartorio": (c.get("nomeCartorio") or c.get("nome") or "").strip(),
            "cnpj": c.get("cnpjCartorio"),
            "cns": c.get("cns"),
            "mensalidade": c.get("valorMensalidade"),
            "totalPago": round(total_pago, 2),
            "totalPendente": round(total_pendente, 2),
            "qtdPaga": qtd_paga,
            "qtdPendente": qtd_pendente,
            "meses": meses_status,
        })

    remessas = [
        {
            "cartorio": l[0],
            "cnpj": l[1],
            "cns": l[2],
            "descricao": l[3],
            "tipo": l[4],
            "valor": l[5],
            "vencimento": l[6],
            "dataPagamento": l[7],
            "status": l[8],
        }
        for l in detalhe_linhas
    ]

    payload = {
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "sources": ["associados.xlsx", "detalhe_remessas.xlsx", "extrato_por_cartorio.xlsx"],
        "meses": meses,
        "associados": cadastros,
        "remessas": remessas,
        "extrato": extrato,
    }

    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    caminho.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"Salvo: {caminho} ({len(cadastros)} associados, {len(remessas)} remessas)")


# ===================== MAIN =====================
def main():
    session = montar_sessao(TOKEN)
    associados_lista = listar_associados(session)
    total = len(associados_lista)

    cadastros = []
    detalhe_linhas = []
    remessas_por_cartorio = defaultdict(dict)
    meses_set = set()
    erros = []

    t0 = time.time()
    for i, a in enumerate(associados_lista, 1):
        aid = a.get("associadoId")
        try:
            detalhe = buscar_detalhe(session, aid)
        except Exception as e:
            log.error(f"  [{aid}] {e}")
            erros.append((aid, a.get("nome"), str(e)))
            continue

        cad = extrair_associado(detalhe)
        cadastros.append(cad)

        for r in detalhe.get("remessas", []):
            descricao = r.get("descricao") or ""
            chave_mes, label_mes = extrair_mes_ano(descricao, r.get("dataVencimento"))
            if chave_mes:
                meses_set.add((chave_mes, label_mes))
                # Se houver duplicidade (ex: "REENVIO"), prioriza a paga
                existente = remessas_por_cartorio[aid].get(chave_mes)
                if existente is None or (not existente.get("dataPago") and r.get("dataPago")):
                    remessas_por_cartorio[aid][chave_mes] = r

            status = "PAGO" if r.get("dataPago") else "PENDENTE"
            detalhe_linhas.append([
                cad["nomeCartorio"] or cad["nome"],
                cad["cnpjCartorio"],
                cad["cns"],
                descricao,
                (r.get("tipoRemessa") or {}).get("tipoRemessaNome"),
                r.get("valorRemessa"),
                fmt_data(r.get("dataVencimento")),
                fmt_data(r.get("dataPago")),
                status,
            ])

        if i % 25 == 0 or i == total:
            elapsed = time.time() - t0
            eta = (elapsed / i) * (total - i)
            log.info(f"  Progresso: {i}/{total}  |  {elapsed:.0f}s decorridos, ETA {eta:.0f}s")

        time.sleep(PAUSA_ENTRE_REQS)

    meses_ordenados = sorted(meses_set, key=lambda x: x[0])

    salvar_cadastro("associados.xlsx", cadastros)
    salvar_detalhe("detalhe_remessas.xlsx", detalhe_linhas)
    salvar_extrato_pivotado("extrato_por_cartorio.xlsx",
                            cadastros, remessas_por_cartorio, meses_ordenados)
    try:
        from export_irtdpj_frontend_data import main as exportar_frontend_data
        exportar_frontend_data()
    except Exception as e:
        log.warning(f"Nao foi possivel gerar o JSON enriquecido do frontend: {e}")
        salvar_frontend_json(FRONTEND_DATA_PATH,
                             cadastros, detalhe_linhas, remessas_por_cartorio, meses_ordenados)

    if erros:
        wb = Workbook()
        ws = wb.active
        ws.title = "Erros"
        ws.append(["associadoId", "nome", "erro"])
        for e in erros:
            ws.append(list(e))
        wb.save("erros.xlsx")
        log.info(f"Salvo: erros.xlsx ({len(erros)} linhas)")

    qtd_pago = sum(1 for l in detalhe_linhas if l[-1] == "PAGO")
    qtd_pend = sum(1 for l in detalhe_linhas if l[-1] == "PENDENTE")
    total_pago = sum((l[5] or 0) for l in detalhe_linhas if l[-1] == "PAGO")
    total_pend = sum((l[5] or 0) for l in detalhe_linhas if l[-1] == "PENDENTE")

    print("\n" + "=" * 64)
    print("  RESUMO")
    print("=" * 64)
    print(f"  Cartórios processados:  {len(cadastros)} / {total}")
    print(f"  Erros:                  {len(erros)}")
    print(f"  Período coberto:        "
          f"{meses_ordenados[0][1] if meses_ordenados else '-'}"
          f" → {meses_ordenados[-1][1] if meses_ordenados else '-'}")
    print(f"  Mensalidades PAGAS:     {qtd_pago:>6}  |  R$ {total_pago:>14,.2f}")
    print(f"  Mensalidades PENDENTES: {qtd_pend:>6}  |  R$ {total_pend:>14,.2f}")
    print("=" * 64)
    print("\nArquivos gerados:")
    print("  • associados.xlsx            (cadastro completo)")
    print("  • detalhe_remessas.xlsx      (1 linha por mensalidade)")
    print("  • extrato_por_cartorio.xlsx  (1 linha por cartório, colunas = meses)")
    print()


if __name__ == "__main__":
    main()
