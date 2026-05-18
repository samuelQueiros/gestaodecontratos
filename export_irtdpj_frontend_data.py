"""
Gera o arquivo consumido pelo frontend a partir das planilhas atuais.

Regra principal:
- status da remessa = status individual do boleto.
- status da competencia = status consolidado por cartorio + mes.
  - se existe qualquer remessa paga no mes, competencia = PAGO.
  - se nao existe paga e existe pendente, competencia = PENDENTE.
  - se nao existe remessa, competencia = SEM_REMESSA.

Uso:
  python export_irtdpj_frontend_data.py
"""

import json
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
OUTPUT = BASE_DIR / "gestaocontratov1" / "public" / "data" / "irtdpj.json"
SOURCES = {
    "associados": BASE_DIR / "associados.xlsx",
    "remessas": BASE_DIR / "detalhe_remessas.xlsx",
    "extrato": BASE_DIR / "extrato_por_cartorio.xlsx",
}


def normalize_key(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def read_sheet(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [
        str(h).strip() if re.fullmatch(r"\d{2}/\d{4}", str(h or "").strip()) else normalize_key(h)
        for h in next(rows)
    ]
    return [dict(zip(headers, row)) for row in rows]


def pick(row, *names):
    for name in names:
        value = row.get(normalize_key(name))
        if value is not None:
            return value
    return None


def only_digits(value):
    return re.sub(r"\D", "", str(value or ""))


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def date_br(value):
    parsed = parse_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else ""


def date_iso(value):
    parsed = parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else None


def month_from(remessa):
    descricao = clean_text(remessa.get("descricao"))
    match = re.search(r"(\d{1,2})/(\d{4})", descricao)
    if match:
        mes = int(match.group(1))
        ano = int(match.group(2))
        if 1 <= mes <= 12:
            return f"{mes:02d}/{ano}", f"{ano}-{mes:02d}"

    vencimento = parse_date(remessa.get("vencimento"))
    if vencimento:
        return vencimento.strftime("%m/%Y"), vencimento.strftime("%Y-%m")

    return "Sem competencia", "9999-99"


def status_from(value):
    return "PAGO" if clean_text(value).upper() == "PAGO" else "PENDENTE"


def normalize_associado(row):
    cnpj_cartorio = pick(row, "CNPJ Cartorio")
    cpf_cnpj = pick(row, "CPF/CNPJ Pessoa")
    nome = clean_text(pick(row, "Nome"))
    cartorio = clean_text(pick(row, "Cartorio")) or nome
    ativo = pick(row, "Ativo")
    return {
        "associadoId": pick(row, "ID"),
        "nome": nome,
        "cpfCnpjPessoa": only_digits(cpf_cnpj),
        "nomeCartorio": cartorio,
        "cnpjCartorio": only_digits(cnpj_cartorio),
        "documentoBusca": only_digits(cnpj_cartorio) or only_digits(cpf_cnpj),
        "cns": clean_text(pick(row, "CNS")) or None,
        "dataAssociacao": date_br(pick(row, "Data Associacao")),
        "dataAssociacaoIso": date_iso(pick(row, "Data Associacao")),
        "valorMensalidade": number(pick(row, "Mensalidade R")),
        "email": clean_text(pick(row, "Email")) or None,
        "telefone": only_digits(pick(row, "Telefone")) or None,
        "ativo": bool(ativo),
        "statusAssociado": "Ativo" if bool(ativo) else "Inativo",
    }


def normalize_remessa(row):
    status = status_from(pick(row, "Status"))
    remessa = {
        "cartorio": clean_text(pick(row, "Cartorio")),
        "cnpj": only_digits(pick(row, "CNPJ")),
        "cns": clean_text(pick(row, "CNS")) or None,
        "descricao": clean_text(pick(row, "Descricao")),
        "tipo": clean_text(pick(row, "Tipo")) or "Nao informado",
        "valor": number(pick(row, "Valor R")),
        "vencimento": date_br(pick(row, "Vencimento")),
        "vencimentoIso": date_iso(pick(row, "Vencimento")),
        "dataPagamento": date_br(pick(row, "Data Pagamento")) or None,
        "dataPagamentoIso": date_iso(pick(row, "Data Pagamento")),
        "status": status,
        "pago": status == "PAGO",
    }
    remessa["mes"], remessa["mesKey"] = month_from(remessa)
    remessa["diasEmAberto"] = 0
    if remessa["status"] == "PENDENTE" and remessa["vencimentoIso"]:
        vencimento = parse_date(remessa["vencimentoIso"])
        if vencimento:
            remessa["diasEmAberto"] = max((datetime.now() - vencimento).days, 0)
    return remessa


def build_competencias(remessas):
    grouped = {}
    for remessa in remessas:
        key = (remessa["cnpj"], remessa["mes"])
        item = grouped.setdefault(
            key,
            {
                "cartorio": remessa["cartorio"],
                "cnpj": remessa["cnpj"],
                "cns": remessa["cns"],
                "mes": remessa["mes"],
                "mesKey": remessa["mesKey"],
                "qtdRemessas": 0,
                "qtdRemessasPagas": 0,
                "qtdRemessasPendentes": 0,
                "valorPagoRemessas": 0.0,
                "valorPendenteRemessas": 0.0,
                "remessas": [],
            },
        )
        item["qtdRemessas"] += 1
        item["remessas"].append(remessa)
        if remessa["status"] == "PAGO":
            item["qtdRemessasPagas"] += 1
            item["valorPagoRemessas"] += remessa["valor"]
        else:
            item["qtdRemessasPendentes"] += 1
            item["valorPendenteRemessas"] += remessa["valor"]

    competencias = []
    for item in grouped.values():
        if item["qtdRemessasPagas"] > 0:
            status_competencia = "PAGO"
            valor_pago = item["valorPagoRemessas"]
            valor_pendente = 0.0
        elif item["qtdRemessasPendentes"] > 0:
            status_competencia = "PENDENTE"
            valor_pago = 0.0
            valor_pendente = item["valorPendenteRemessas"]
        else:
            status_competencia = "SEM_REMESSA"
            valor_pago = 0.0
            valor_pendente = 0.0

        competencia = {
            **{k: v for k, v in item.items() if k != "remessas"},
            "statusCompetencia": status_competencia,
            "valorPago": round(valor_pago, 2),
            "valorPendente": round(valor_pendente, 2),
            "valorPagoRemessas": round(item["valorPagoRemessas"], 2),
            "valorPendenteRemessas": round(item["valorPendenteRemessas"], 2),
            "temBoletoPendenteIgnorado": status_competencia == "PAGO"
            and item["qtdRemessasPendentes"] > 0,
        }
        competencias.append(competencia)

        for remessa in item["remessas"]:
            remessa["statusCompetencia"] = status_competencia
            remessa["competenciaConsolidadaPaga"] = status_competencia == "PAGO"
            remessa["pendenciaIgnoradaPorPagamentoNaCompetencia"] = (
                remessa["status"] == "PENDENTE" and status_competencia == "PAGO"
            )

    return sorted(competencias, key=lambda item: (item["mesKey"], item["cartorio"]))


def normalize_extrato(row, meses, competencias_por_chave):
    cnpj = only_digits(pick(row, "CNPJ"))
    status_meses = {}
    for mes in meses:
        competencia = competencias_por_chave.get((cnpj, mes))
        raw_text = clean_text(row.get(normalize_key(mes)))
        if competencia:
            status = competencia["statusCompetencia"]
            if status == "PAGO":
                texto = raw_text if raw_text.startswith("PAGO") else "PAGO"
            elif status == "PENDENTE":
                texto = "PENDENTE"
            else:
                texto = ""
        else:
            status = "SEM_REMESSA"
            texto = ""

        status_meses[mes] = {
            "texto": texto,
            "status": status,
            "qtdRemessas": competencia["qtdRemessas"] if competencia else 0,
            "qtdRemessasPagas": competencia["qtdRemessasPagas"] if competencia else 0,
            "qtdRemessasPendentes": competencia["qtdRemessasPendentes"] if competencia else 0,
            "valorPago": competencia["valorPago"] if competencia else 0,
            "valorPendente": competencia["valorPendente"] if competencia else 0,
        }

    total_pago = sum(item["valorPago"] for item in status_meses.values())
    total_pendente = sum(item["valorPendente"] for item in status_meses.values())
    qtd_paga = sum(1 for item in status_meses.values() if item["status"] == "PAGO")
    qtd_pendente = sum(1 for item in status_meses.values() if item["status"] == "PENDENTE")

    return {
        "cartorio": clean_text(pick(row, "Cartorio")),
        "cnpj": cnpj,
        "cns": clean_text(pick(row, "CNS")) or None,
        "mensalidade": number(pick(row, "Mensalidade R")),
        "totalPago": round(total_pago, 2),
        "totalPendente": round(total_pendente, 2),
        "qtdPaga": qtd_paga,
        "qtdPendente": qtd_pendente,
        "totalPagoRemessas": number(pick(row, "Total Pago R")),
        "totalPendenteRemessas": number(pick(row, "Total Pendente R")),
        "meses": {mes: status_meses[mes]["texto"] for mes in meses},
        "mesesDetalhados": status_meses,
        "temPendencia": any(item["status"] == "PENDENTE" for item in status_meses.values()),
    }


def main():
    associados_rows = read_sheet(SOURCES["associados"])
    remessas_rows = read_sheet(SOURCES["remessas"])
    extrato_rows = read_sheet(SOURCES["extrato"])

    meses = [
        key
        for key in extrato_rows[0].keys()
        if isinstance(key, str) and re.fullmatch(r"\d{2}/\d{4}", key)
    ] if extrato_rows else []
    meses = sorted(meses, key=lambda label: (int(label[3:]), int(label[:2])))

    associados = [normalize_associado(row) for row in associados_rows]
    remessas = [normalize_remessa(row) for row in remessas_rows]
    competencias = build_competencias(remessas)
    competencias_por_chave = {(item["cnpj"], item["mes"]): item for item in competencias}
    extrato = [normalize_extrato(row, meses, competencias_por_chave) for row in extrato_rows]

    tipos = sorted({remessa["tipo"] for remessa in remessas if remessa["tipo"]})
    cartorios = sorted({item["cartorio"] for item in extrato if item["cartorio"]})

    payload = {
        "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "sources": [source.name for source in SOURCES.values()],
        "meta": {
            "totalAssociados": len(associados),
            "totalRemessas": len(remessas),
            "totalCompetencias": len(competencias),
            "totalExtrato": len(extrato),
            "tipos": tipos,
            "cartorios": cartorios,
        },
        "meses": meses,
        "associados": associados,
        "remessas": remessas,
        "competencias": competencias,
        "extrato": extrato,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Arquivo gerado: {OUTPUT}")
    print(
        f"Associados: {len(associados)} | Remessas: {len(remessas)} | "
        f"Competencias: {len(competencias)} | Extrato: {len(extrato)}"
    )


if __name__ == "__main__":
    main()
